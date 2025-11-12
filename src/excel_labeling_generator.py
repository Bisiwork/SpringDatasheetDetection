
"""
excel_labeling_generator.py - Generatore workbook Excel per labeling manuale.

Questo modulo crea workbook Excel strutturati per il labeling manuale dei parametri
estratti da datasheet di molle. Per ciascun file di input:
- Crea un foglio dedicato con colonne: parameter, predicted, ground_truth
- Inserisce immagini/PDF renderizzati come riferimento visivo
- Applica styling professionale (header, colori, allineamenti)
- Supporta priorità campi e ordinamento personalizzato

Input: Immagini/PDF in input_dir, predizioni JSON in outputs_dir
Output: Workbook Excel in workbook_path con fogli per ogni file + README
Dipendenze: openpyxl, PIL, pdf2image (opzionale)
"""
from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any, Dict, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

try:
    from pdf2image import convert_from_path
except Exception:
    convert_from_path = None  # opzionale: attiva se installi pdf2image+poppler

# Config
IMG_EXTS = {".jpg", ".jpeg", ".png", ".gif"}
PDF_EXTS = {".pdf"}


def _load_json(p: Path) -> Dict[str, Any]:
    return json.loads(p.read_text(encoding="utf-8"))


def _fit_image_for_excel(img_path: Path, max_px: tuple[int, int] = (900, 900)) -> Path:
    """
    Ridimensiona (thumbnail) l’immagine per inserirla in Excel, preservando l’aspect ratio.
    Ritorna un path temporaneo PNG senza creare file .thumb.png ridondanti.
    Usa PIL.Image.thumbnail e salva in memoria o file temporaneo.
    """
    import tempfile
    import os

    im = PILImage.open(img_path)
    im = im.convert("RGB")
    im.thumbnail(max_px)
    # Usa file temporaneo invece di .thumb.png
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        im.save(tmp.name, format="PNG", optimize=True)
        return Path(tmp.name)


def _render_pdf_first_page(pdf_path: Path, dpi: int = 180) -> Path:
    """
    Rende la prima pagina del PDF a immagine per Excel.
    Richiede pdf2image + poppler installati nel sistema (su mac: brew install poppler).
    """

    if convert_from_path is None:
        raise RuntimeError("pdf2image non disponibile. Installa 'pdf2image' e Poppler.")
    pages = convert_from_path(str(pdf_path), dpi=dpi)
    if not pages:
        raise RuntimeError(f"Nessuna pagina renderizzata per {pdf_path}")
    out = pdf_path.with_suffix(".page1.png")
    pages[0].save(out, "PNG")
    return out


def _safe_sheet_name(name: str) -> str:
    s = "".join(ch if ch not in '[]:*?/\\' else "_" for ch in name)
    return s[:31]


def build_labeling_workbook(
    input_dir: Path | str = "data/all",
    outputs_dir: Path | str = "output/exp1/results/<MODEL>/outputs",
    workbook_path: Path | str = "output/exp1/reports/<MODEL>_labeling.xlsx",
    *,
    fields_priority: Sequence[str] | None = None,
    ground_truth_dir: Path | str | None = None,
):
    # Mapping from input stem to ground truth stem
    gt_stem_mapping = {
        'Fertigungsversuch 0.18mm Simplex MX5.page1': 'Fertigungsversuch 0.18mm Simple',
        'Soluzione2_FiloRettBlocco46,4.page1': 'Soluzione2_FiloRettBlocco46,4.p',
        'Soluzione3_FiloRettCorsaInalterata.page1': 'Soluzione3_FiloRettCorsaInalter',
        'Spring 2 - wire diameter 0,23.page1': 'Spring 2 - wire diameter 0,23.p',
        'Spring 1 - wire diameter 0,20.page1': 'Spring 1 - wire diameter 0,20.p',
        'P294090_B_38-12_Tri_Wedge_Installation_Spring.page1': 'P294090_B_38-12_Tri_Wedge_Insta',
        'Fertigungsversuch 0.16mm Simplex MX5.page1': 'Fertigungsversuch 0.16mm Simple',
        '34083594A - A0033C2168.cgm.page1': '34083594A - A0033C2168.cgm.page',
        'Zeicunungen vom Baldim[2].page1': 'Zeicunungen vom Baldim_2_.page1',
    }
    """
    Per ogni file in input_dir, crea un foglio:
      A: 'parameter'
      B: 'predicted'
      C: 'ground_truth' (vuoto da compilare)
      immagine/screenshot ancorata in E2.
    Cerca predizioni in outputs_dir/<stem>.json (se assente, colonna B vuota).
    """

    input_dir = Path(input_dir)
    outputs_dir = Path(outputs_dir)
    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    ground_truth_dir = Path(ground_truth_dir) if ground_truth_dir else None

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README"
    ws0["A1"] = "Istruzioni"
    ws0["A1"].font = Font(bold=True, size=24)
    ws0["A2"] = "Compila la colonna C ('ground_truth') con i valori corretti per ogni parametro. L'immagine a destra è un riferimento."
    ws0["A2"].font = Font(bold=True, size=18)

    temp_files = []  # Keep track of temporary files to delete after saving
    for f in sorted([p for p in input_dir.iterdir() if p.suffix.lower() in (IMG_EXTS | PDF_EXTS) and "thumb" not in p.stem]):
        stem = f.stem
        pred_path = outputs_dir / f"{stem}.json"
        preds = _load_json(pred_path) if pred_path.exists() else {}
        # If not found, try with .page1.json
        if not preds:
            pred_path_page1 = outputs_dir / f"{stem}.page1.json"
            preds = _load_json(pred_path_page1) if pred_path_page1.exists() else {}
        # Remove debugTrail and usage as they're not needed for labeling
        preds.pop('debugTrail', None)
        preds.pop('usage', None)

        # Load ground truth if available
        gt = {}
        if ground_truth_dir:
            gt_stem = gt_stem_mapping.get(stem, stem)
            gt_path = ground_truth_dir / f"{gt_stem}.json"
            if gt_path.exists():
                gt = _load_json(gt_path)

        keys = list(preds.keys())
        if fields_priority:
            front = [k for k in fields_priority if k in keys]
            rest = sorted([k for k in keys if k not in fields_priority])
            keys = front + rest
        else:
            keys = sorted(keys)

        ws = wb.create_sheet(_safe_sheet_name(stem))
        headers = ["parameter", "predicted", "ground_truth"]
        ws.append(headers)
        for k in keys:
            ws.append([k, preds.get(k, ""), gt.get(k, "")])

        # Style headers
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=18, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style parameter names (column A)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Style predicted values (column B)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            cell.font = Font(size=14, color="1F497D")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Style ground truth values (column C)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            cell.font = Font(size=14)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.freeze_panes = "A2"

        try:
            if f.suffix.lower() in IMG_EXTS:
                img_path = _fit_image_for_excel(f)
            else:
                img_png = _render_pdf_first_page(f)
                img_path = _fit_image_for_excel(img_png)
            temp_files.append(img_path)  # Track for later deletion
            xl_img = XLImage(str(img_path))
            ws.add_image(xl_img, "E2")
        except Exception as e:
            ws["E2"] = f"Immagine non disponibile: {e}"

    wb.save(str(workbook_path))
    # Clean up temporary files after saving
    for temp_file in temp_files:
        temp_file.unlink(missing_ok=True)
    return workbook_path
