
from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any, Dict, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

try:
    from pdf2image import convert_from_path
except Exception:
    convert_from_path = None  # opzionale: attiva se installi pdf2image+poppler

# Config
IMG_EXTS = {".jpg", ".jpeg", ".png", ".gif"}
PDF_EXTS = {".pdf"}


def _load_json(p: Path) -> Dict[str, Any]:
    return json.loads(p.read_text(encoding="utf-8"))


def _fit_image_for_excel(img_path: Path, max_px: tuple[int, int] = (900, 900)) -> Path:
    """
    Ridimensiona (thumbnail) l’immagine per inserirla in Excel, preservando l’aspect ratio.
    Ritorna un path temporaneo PNG. Usa PIL.Image.thumbnail.
    """

    im = PILImage.open(img_path)
    im = im.convert("RGB")
    im.thumbnail(max_px)
    out = img_path.with_suffix(".thumb.png")
    im.save(out, format="PNG", optimize=True)
    return out


def _render_pdf_first_page(pdf_path: Path, dpi: int = 180) -> Path:
    """
    Rende la prima pagina del PDF a immagine per Excel.
    Richiede pdf2image + poppler installati nel sistema (su mac: brew install poppler).
    """

    if convert_from_path is None:
        raise RuntimeError("pdf2image non disponibile. Installa 'pdf2image' e Poppler.")
    pages = convert_from_path(str(pdf_path), dpi=dpi)
    if not pages:
        raise RuntimeError(f"Nessuna pagina renderizzata per {pdf_path}")
    out = pdf_path.with_suffix(".page1.png")
    pages[0].save(out, "PNG")
    return out


def _safe_sheet_name(name: str) -> str:
    s = "".join(ch if ch not in '[]:*?/\\' else "_" for ch in name)
    return s[:31]


def build_labeling_workbook(
    input_dir: Path | str = "data/sketch",
    outputs_dir: Path | str = "results/<MODEL>/outputs",
    workbook_path: Path | str = "reports/<MODEL>_labeling.xlsx",
    *,
    fields_priority: Sequence[str] | None = None,
):
    """
    Per ogni file in input_dir, crea un foglio:
      A: 'parameter'
      B: 'predicted'
      C: 'ground_truth' (vuoto da compilare)
      immagine/screenshot ancorata in E2.
    Cerca predizioni in outputs_dir/<stem>.json (se assente, colonna B vuota).
    """

    input_dir = Path(input_dir)
    outputs_dir = Path(outputs_dir)
    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README"
    ws0["A1"] = "Istruzioni"
    ws0["A1"].font = Font(bold=True, size=24)
    ws0["A2"] = "Compila la colonna C ('ground_truth') con i valori corretti per ogni parametro. L'immagine a destra è un riferimento."
    ws0["A2"].font = Font(bold=True, size=18)

    for f in sorted([p for p in input_dir.iterdir() if p.suffix.lower() in (IMG_EXTS | PDF_EXTS) and "thumb" not in p.stem]):
        stem = f.stem
        pred_path = outputs_dir / f"{stem}.json"
        preds = _load_json(pred_path) if pred_path.exists() else {}

        keys = list(preds.keys())
        if fields_priority:
            front = [k for k in fields_priority if k in preds]
            rest = sorted([k for k in keys if k not in fields_priority])
            keys = front + rest
        else:
            keys = sorted(keys)

        ws = wb.create_sheet(_safe_sheet_name(stem))
        headers = ["parameter", "predicted", "ground_truth"]
        ws.append(headers)
        for k in keys:
            ws.append([k, preds.get(k, ""), ""])

        # Style headers
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=18, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style parameter names (column A)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Style predicted values (column B)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            cell.font = Font(size=14, color="1F497D")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Style ground truth values (column C)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            cell.font = Font(size=14)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.freeze_panes = "A2"

        try:
            if f.suffix.lower() in IMG_EXTS:
                img_path = _fit_image_for_excel(f)
            else:
                img_png = _render_pdf_first_page(f)
                img_path = _fit_image_for_excel(img_png)
            xl_img = XLImage(str(img_path))
            ws.add_image(xl_img, "E2")
        except Exception as e:
            ws["E2"] = f"Immagine non disponibile: {e}"

    wb.save(str(workbook_path))
    return workbook_path
